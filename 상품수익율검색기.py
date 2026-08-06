# -*- coding: utf-8 -*-
"""
상품 수익율 검색기 (Streamlit)

데이터: 스크립트와 같은 폴더의 "NN년도매출.parquet" (권장) 또는 "NN년도매출.xlsx"
        → xlsx는 로컬에서 '변환_xlsx_to_parquet.py'로 미리 변환해서 올릴 것

산식
  수익율 = 수익원(실배송비) / 최종판매가 * 100
  정산금 = 출고원가 + 수익원(실배송비)   (= 판매가 - 수수료)

정제
  - 반품(음수 수량) 행은 원매출건과 함께 제거 (주문번호+모델명 키 단위)
  - 판매가 0원 행(쇼핑백/사은품 등) 제외

실행: streamlit run 상품수익율검색기.py
"""

import glob
import hashlib
import hmac
import os
import re

import altair as alt
import numpy as np
import pandas as pd
import streamlit as st

# ──────────────────────────────────────────────
# 설정
# ──────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 데이터는 저장소 루트든 하위 폴더(data/ 등)든 어디에 있어도 찾는다.
# 파일명도 한글·영문 상관없다 (매출 여부는 파일명이 아니라 컬럼 구성으로 판별).
SEARCH_GLOBS = ("*.{ext}", "*/*.{ext}")
MALL_CLASS_PATTERN = os.path.join(BASE_DIR, "*몰분류*.xlsx")   # 분류=매장 → 통계 제외
STOCK_NAME = "재고.parquet"                                     # 변환기가 생성
CACHE_DIR = os.path.join(BASE_DIR, "_cache")


def _찾기(확장자):
    """루트와 바로 아래 하위 폴더에서 해당 확장자 파일을 모은다."""
    나온것 = []
    for 틀 in SEARCH_GLOBS:
        for f in glob.glob(os.path.join(BASE_DIR, 틀.format(ext=확장자))):
            if os.path.basename(f).startswith("~$"):
                continue
            if os.path.abspath(os.path.dirname(f)) == os.path.abspath(CACHE_DIR):
                continue                                  # 변환 캐시는 제외
            나온것.append(f)
    return sorted(set(나온것))

HEADER_ROW = 1   # 헤더가 2번째 행
SHEET_NAME = 0   # 첫 번째 시트
XLSX_SIZE_LIMIT_MB = 20   # 이보다 큰 xlsx는 서버에서 변환하지 않음 (메모리 초과 방지)

COL_ORDER = "주문번호"
COL_MODEL = "모델명"
COL_MALL = "쇼핑몰"
COL_BRAND = "브랜드"
COL_QTY = "수량"
COL_COST = "출고원가"
COL_PRICE = "최종판매가"
COL_PROFIT = "수익원(실배송비)"
COL_DATE = "출고날짜"
COL_NOTE = "비고"
COL_FEE = "수수료액"          # 새 변환기부터 담긴다 (없으면 예전 방식으로 동작)
COL_SHIP = "실배송비(품목별)"

USE_COLS = [COL_ORDER, COL_MALL, COL_BRAND, "대카테고리", "카테고리",
            COL_MODEL, COL_QTY, COL_COST, COL_PRICE, COL_FEE, COL_SHIP,
            COL_PROFIT, COL_DATE, COL_NOTE]

CAT_COLS = [COL_MALL, COL_BRAND, "대카테고리", "카테고리", COL_NOTE, COL_MODEL]

st.set_page_config(page_title="상품 수익율 검색기", page_icon="🔍", layout="wide")


# ──────────────────────────────────────────────
# 비밀번호 잠금
# ──────────────────────────────────────────────
# 평문 대신 SHA-256 해시만 보관 (저장소에 비밀번호가 그대로 남지 않도록)
PASSWORD_SHA256 = "c356b589ca5af32ef0049f110a6c82d03f78b7a791bc85eda9e1793b72bd86ee"


def _password_ok(pw: str) -> bool:
    """Secrets에 app_password가 있으면 그 값을, 없으면 내장 해시를 사용"""
    try:
        expected = st.secrets.get("app_password")
    except Exception:
        expected = None
    if expected:
        return hmac.compare_digest(pw, str(expected))
    return hmac.compare_digest(
        hashlib.sha256(pw.encode("utf-8")).hexdigest(), PASSWORD_SHA256)


def require_login() -> None:
    if st.session_state.get("_authed"):
        return

    st.title("🔒 상품 수익율 검색기")
    st.caption("비밀번호를 입력하세요.")
    with st.form("login_form"):
        pw = st.text_input("비밀번호", type="password", label_visibility="collapsed",
                           placeholder="비밀번호")
        submitted = st.form_submit_button("로그인")

    if submitted:
        if _password_ok(pw):
            st.session_state["_authed"] = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")
    st.stop()


require_login()

st.markdown("""<style>
div[data-testid="stMetricValue"] { font-size: 1.55rem; }
div[data-testid="stMetricLabel"] { font-size: 0.78rem; }
h1 { font-size: 1.6rem !important; }
h3 { font-size: 1.15rem !important; }
</style>""", unsafe_allow_html=True)


# ──────────────────────────────────────────────
# 데이터 로딩
# ──────────────────────────────────────────────
def _재고파일():
    """재고.parquet 을 루트·하위 폴더에서 찾는다. 여러 개면 가장 최근 것."""
    후보 = [f for f in _찾기("parquet") if os.path.basename(f) == STOCK_NAME]
    if not 후보:
        return None
    return max(후보, key=os.path.getmtime)


@st.cache_data(show_spinner=False)
def load_stock(sig):
    if sig is None:
        return None
    df = pd.read_parquet(sig[0])
    df["모델명_U"] = df["모델명"].astype(str).str.upper()
    return df


def get_stock_sig():
    경로 = _재고파일()
    if 경로 is None:
        return None
    return (경로, int(os.path.getmtime(경로)))


def _이미지파일():
    """모델명 → 이미지 URL 표를 찾는다. parquet 우선(가볍고 빠름), 없으면 xlsx.

    저장소에는 .gitignore 가 xlsx 를 막고 있어 parquet 만 올라간다.
    로컬에서 이미지.xlsx 만 갱신했을 때도 그대로 뜨도록 xlsx 도 받아준다.
    """
    후보 = [f for 확장자 in ("parquet", "xlsx")
           for f in _찾기(확장자) if "이미지" in os.path.basename(f)]
    if not 후보:
        return None
    # 더 최근에 갱신된 쪽을 쓴다. xlsx 를 새로 받아 놓고 parquet 을 안 만들었어도
    # 바로 반영되도록. (같은 시각이면 가벼운 parquet 우선)
    return max(후보, key=lambda f: (os.path.getmtime(f), f.lower().endswith(".parquet")))


def get_image_sig():
    경로 = _이미지파일()
    if 경로 is None:
        return None
    return (경로, int(os.path.getmtime(경로)))


@st.cache_data(show_spinner=False)
def load_images(sig):
    if sig is None:
        return None
    경로 = sig[0]
    try:
        if 경로.lower().endswith(".parquet"):
            m = pd.read_parquet(경로)
        else:
            # dtype=str: 모델명이 숫자로만 되어 있으면 590728 이 590728.0 으로
            # 읽혀 매칭이 통째로 어긋난다. 앞자리 0 도 날아간다.
            m = pd.read_excel(경로, sheet_name="이미지", dtype=str)
        m.columns = [str(c).strip() for c in m.columns]
        if "모델명" not in m.columns or "이미지" not in m.columns:
            return None
        m = m[["모델명", "이미지"]].dropna()
        키 = m["모델명"].astype(str).str.strip().str.upper()
        return dict(zip(키, m["이미지"].astype(str).str.strip()))
    except Exception:
        return None


def _이미지주소(url: str) -> str:
    """URL에 공백이 들어 있어 (예: '.../A1050186 868.jpg') 그대로 쓰면 깨질 수 있다."""
    from urllib.parse import quote
    return quote(url, safe=":/?&=%#")


_RX_SIZE = re.compile(r"\s*\([^)]*\)\s*$")


def _이미지키들(이름):
    """이미지 표를 찾을 때 시도할 이름들.

    이미지 표는 사이즈가 빠진 '라인명' 기준이다.
    (예: 표에는 '10AW220 8507', 재고·매출에는 '10AW220 8507 (40)')
    그래서 원래 이름 → 사이즈 괄호를 뗀 이름 순으로 찾는다.
    """
    if 이름 is None or (isinstance(이름, float) and pd.isna(이름)):
        return []
    원본 = str(이름).strip().upper()
    if not 원본:
        return []
    벗김 = _RX_SIZE.sub("", 원본).strip()
    return [원본] if 벗김 == 원본 else [원본, 벗김]


def _이미지찾기(img_map, 후보들):
    if not img_map:
        return None
    본것 = set()
    for c in 후보들:
        for k in _이미지키들(c):
            if k in 본것:
                continue
            본것.add(k)
            u = img_map.get(k)
            if u:
                return u
    return None


def get_mall_class_sig():
    files = sorted(glob.glob(MALL_CLASS_PATTERN))
    if not files:
        return None
    return (files[0], int(os.path.getmtime(files[0])))


def load_store_malls(sig) -> frozenset:
    """몰분류 파일에서 분류='매장'인 몰 목록"""
    if sig is None:
        return frozenset()
    m = pd.read_excel(sig[0], header=0)
    m = m.iloc[:, :2]
    m.columns = ["몰명", "분류"]
    return frozenset(m.loc[m["분류"].astype(str).str.strip() == "매장", "몰명"]
                     .dropna().astype(str).str.strip())


def _is_sales_parquet(path: str) -> bool:
    """스키마에 주문번호·출고날짜가 있으면 매출 데이터로 인식"""
    try:
        import pyarrow.parquet as pq
        cols = set(pq.ParquetFile(path).schema_arrow.names)
        return {"주문번호", "출고날짜", "쇼핑몰"} <= cols
    except Exception:
        return False


def get_file_sigs():
    # parquet: 파일명·폴더 무관, 스키마로 매출 데이터 판별 (재고.parquet 등은 자동 제외)
    pq_files = [f for f in _찾기("parquet") if _is_sales_parquet(f)]
    pq_bases = {os.path.splitext(os.path.basename(f))[0] for f in pq_files}
    xlsx_files = [f for f in _찾기("xlsx")
                  if "매출" in os.path.basename(f)
                  and os.path.splitext(os.path.basename(f))[0] not in pq_bases]
    files = pq_files + xlsx_files
    return tuple((f, int(os.path.getmtime(f))) for f in files)


def _read_one(path: str) -> pd.DataFrame:
    """parquet은 바로, xlsx는 캐시 변환 후 로딩"""
    if path.lower().endswith(".parquet"):
        return pd.read_parquet(path)

    size_mb = os.path.getsize(path) / 1024 / 1024
    if size_mb > XLSX_SIZE_LIMIT_MB:
        raise MemoryError(
            f"'{os.path.basename(path)}'({size_mb:.0f}MB)는 서버에서 직접 변환하기에 너무 큽니다.\n"
            f"로컬 PC에서 '변환_xlsx_to_parquet.py'를 실행해 .parquet으로 만든 뒤 그 파일만 올려주세요."
        )

    os.makedirs(CACHE_DIR, exist_ok=True)
    base = os.path.splitext(os.path.basename(path))[0]
    pq = os.path.join(CACHE_DIR, f"{base}_{int(os.path.getmtime(path))}.parquet")
    if os.path.exists(pq):
        return pd.read_parquet(pq)

    with st.spinner(f"{os.path.basename(path)} 변환 중... (최초 1회)"):
        df = pd.read_excel(path, sheet_name=SHEET_NAME, header=HEADER_ROW)
        df = df[[c for c in USE_COLS if c in df.columns]]
        df.to_parquet(pq)
    for old in glob.glob(os.path.join(CACHE_DIR, f"{base}_*.parquet")):
        if old != pq:
            try:
                os.remove(old)
            except OSError:
                pass
    return df


def _키해시(df: pd.DataFrame):
    """주문번호+모델명을 64비트 숫자로. (문자열을 그대로 들고 있으면 메모리를 크게 먹는다)"""
    키 = df[COL_ORDER].astype(str) + "\x00" + df[COL_MODEL].astype(str)
    return pd.util.hash_pandas_object(키, index=False).to_numpy()


@st.cache_data(show_spinner="데이터 불러오는 중...")
def load_all_data(file_sigs: tuple, mall_sig=None) -> pd.DataFrame:
    # 파일끼리 겹치는 구간이 있다 (연도 파일의 자투리, 월별 파일과 지난 스냅샷 등).
    # 그냥 이어붙이면 같은 주문이 두 번 잡혀 매출이 부풀려진다.
    # → 최근에 갱신된 파일부터 읽고, 앞서 나온 주문번호+모델명은 버린다.
    #   (같은 파일 안의 중복은 정상 데이터일 수 있으므로 건드리지 않는다)
    순서 = sorted(file_sigs, key=lambda x: x[1], reverse=True)

    frames, 본키 = [], np.empty(0, dtype=np.uint64)
    버린수 = 0
    for 경로, _ in 순서:
        조각 = _read_one(경로)
        if COL_ORDER in 조각.columns and COL_MODEL in 조각.columns:
            해시 = _키해시(조각)
            겹침 = np.isin(해시, 본키)
            if 겹침.any():
                버린수 += int(겹침.sum())
                조각 = 조각[~겹침]
                해시 = 해시[~겹침]
            본키 = np.union1d(본키, 해시)
        frames.append(조각)

    df = pd.concat(frames, ignore_index=True)
    del frames, 본키

    # 필수 컬럼 확인
    missing = [c for c in (COL_ORDER, COL_MODEL, COL_QTY, COL_COST,
                           COL_PRICE, COL_PROFIT, COL_DATE) if c not in df.columns]
    if missing:
        raise KeyError(f"필수 컬럼 누락: {missing} — 변환기 버전을 확인하세요.")

    # 날짜
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors="coerce")
    df = df.dropna(subset=[COL_DATE])

    # 숫자형
    for c in (COL_QTY, COL_COST, COL_PRICE, COL_PROFIT):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # 반품 제거: 음수 수량이 포함된 (주문번호+모델명) 키 전체 삭제
    key = pd.factorize(df[COL_ORDER].astype(str) + "\x00" + df[COL_MODEL].astype(str))[0]
    bad = np.unique(key[df[COL_QTY].values < 0])
    df = df[~np.isin(key, bad)]
    del key, bad

    # 판매가 0원 행 제외
    df = df[df[COL_PRICE] > 0].drop(columns=[COL_ORDER])

    # 매장(오프라인) 판매는 지우지 않고 표시만 해 둔다.
    # 판매수량·판매금액·로우데이터에는 포함하고,
    # 수익율·이익율·등급을 낼 때만 빼기 위해서다.
    store_malls = load_store_malls(mall_sig)
    df["매장"] = (df[COL_MALL].astype(str).str.strip().isin(store_malls)
                if store_malls else False)

    # 메모리 절감
    for c in CAT_COLS:
        if c in df.columns:
            df[c] = df[c].astype("category")
    df[COL_QTY] = df[COL_QTY].astype("int32")
    for c in (COL_COST, COL_PRICE, COL_PROFIT):
        df[c] = df[c].astype("float32")

    # 파생
    df["연도"] = df[COL_DATE].dt.year.astype("int16")
    df["수익율"] = (df[COL_PROFIT] / df[COL_PRICE] * 100).astype("float32")

    # 정산금은 두 가지로 나눈다.
    #   정산금       = 최종판매가 - 수수료액        (실제로 정산돼 들어온 돈. 배송비 차감 안 함)
    #   정산금_수익   = 정산금 - 실배송비            (= 출고원가 + 수익원(실배송비))
    # 화면에 보여주는 금액은 앞의 것, 수익율/이익율 계산은 뒤의 것을 쓴다.
    df["정산금_수익"] = (df[COL_COST] + df[COL_PROFIT]).astype("float32")
    if COL_FEE in df.columns:
        # 수수료액이 있는 행만 새 산식. 옛 parquet(24/25년 등)에는 그 값이 없는데
        # 0 으로 채우면 수수료를 안 뺀 판매가가 그대로 정산금이 되어 크게 부풀려진다.
        수수료 = pd.to_numeric(df[COL_FEE], errors="coerce")
        df["정산금"] = np.where(수수료.notna(),
                             df[COL_PRICE] -수수료.fillna(0),
                             df["정산금_수익"]).astype("float32")
    else:
        # 수수료액 컬럼 자체가 없으면 예전 방식 그대로
        df["정산금"] = df["정산금_수익"]

    out = df.reset_index(drop=True)
    out.attrs["중복제거"] = 버린수
    return out


# ──────────────────────────────────────────────
# 포맷 헬퍼
# ──────────────────────────────────────────────
def fmt_won(v):
    return "-" if v is None or pd.isna(v) else f"{v:,.0f}원"


def fmt_pct(v):
    return "-" if v is None or pd.isna(v) else f"{v:.2f}%"


WON = st.column_config.NumberColumn(format="localized")
NUM = st.column_config.NumberColumn(format="localized")

# 코드성 컬럼. 엑셀에서 '텍스트' 서식으로 넣어야 값이 안 망가진다.
#   710548524014 → 7.10548E+11 로 바뀌고, 앞자리 0 은 그냥 사라진다.
TEXT_COLS = ("라인명", "브랜드", "대카테고리", "카테고리", "종류", "성별", "모델명",
             "등급", "기준기간")


@st.cache_data(show_spinner=False)
def _엑셀바이트(df: pd.DataFrame, 시트명: str) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 시트명

    ws.append(list(df.columns))
    for 칸 in ws[1]:
        칸.font = Font(bold=True)
        칸.alignment = Alignment(horizontal="center")

    for 행 in df.itertuples(index=False):
        ws.append(["" if pd.isna(v) else v for v in 행])

    최대너비 = {"모델명": 60, "라인명": 28}
    for i, 이름 in enumerate(df.columns, start=1):
        문자 = get_column_letter(i)
        if 이름 in TEXT_COLS:
            for 칸 in ws[문자][1:]:          # 머리글 제외
                칸.number_format = "@"       # 텍스트
        elif 이름.endswith("(%)"):
            for 칸 in ws[문자][1:]:
                칸.number_format = "0.00"
        else:
            for 칸 in ws[문자][1:]:
                칸.number_format = "#,##0"
        길이 = max([len(str(이름))] + [len(str(v)) for v in df[이름].head(300)])
        ws.column_dimensions[문자].width = min(길이 + 2, 최대너비.get(이름, 20))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _행사가표시(info) -> str:
    """재고 파일에 들어 있는 EC행사가·수익률을 재고 칸 아래에 덧붙인다."""
    가 = info.get("EC행사가")
    율 = info.get("EC행사수익률")
    자 = info.get("자사몰행사가")
    if 가 is None or pd.isna(가) or 가 <= 0:
        return ""
    조각 = [f"EC행사가 <b>{가:,.0f}원</b>"]
    if 자 is not None and pd.notna(자) and 자 > 0:
        조각.append(f"자사몰 <b>{자:,.0f}원</b>")
    return ("<br><span style='border-top:1px dashed #e6e6e6;display:block;"
            "margin-top:0.3rem;padding-top:0.3rem'>" + " · ".join(조각) + "</span>")


def _한줄지표(라벨1, 값1, 라벨2, 값2):
    """좁은 칸에 지표 두 개를 한 줄로. st.metric 보다 작은 글씨로 찍는다."""
    st.markdown(
        "<div style='display:flex;gap:0.9rem;align-items:baseline;flex-wrap:wrap;"
        "margin:0.15rem 0 0.6rem'>"
        + "".join(
            f"<div><div style='font-size:0.72rem;color:#666;line-height:1.2'>{라벨}</div>"
            f"<div style='font-size:1.02rem;font-weight:600;line-height:1.25'>{값}</div></div>"
            for 라벨, 값 in ((라벨1, 값1), (라벨2, 값2)))
        + "</div>",
        unsafe_allow_html=True)


# ──────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────
st.title("🔍 상품 수익율 검색기")

file_sigs = get_file_sigs()
if not file_sigs:
    st.error("데이터 파일이 없습니다. 이 스크립트와 같은 폴더에 "
             "파일명에 '매출'이 들어간 .parquet(또는 .xlsx)을 넣어주세요.")
    st.caption(f"현재 폴더: {BASE_DIR}")
    st.stop()

mall_sig = get_mall_class_sig()

try:
    df = load_all_data(file_sigs, mall_sig)
except Exception as e:
    st.error(f"데이터 로딩 실패: {e}")
    st.exception(e)
    st.stop()

st.markdown(
    "<div style='font-size:0.82rem;color:#555;line-height:1.75;"
    "background:#f6f8fa;border:1px solid #e6e8eb;border-radius:6px;"
    "padding:0.5rem 0.75rem;margin-bottom:0.5rem'>"
    "<b>상품등급 산정 규칙</b> — 이익율(수익 ÷ 정산금) 기준<br>"
    "<span style='color:#7c3aed;font-weight:700'>S</span> 40% 이상 <b>이면서</b> 판매 20개 이상 · "
    "<span style='color:#1a7f37;font-weight:700'>A</span> 25% 이상 · "
    "<span style='color:#0969da;font-weight:700'>B</span> 15% 이상 · "
    "<span style='color:#bf8700;font-weight:700'>C</span> 5% 이상 · "
    "<span style='color:#d4691e;font-weight:700'>D</span> −5% 이상 · "
    "<span style='color:#cf222e;font-weight:700'>E</span> −5% 미만<br>"
    "· 기간은 최근 3개월부터, 100건을 넘길 때까지 3개월씩 넓힘 (최대 24개월)<br>"
    "· 매장(오프라인) 판매는 등급·수익율에서 제외 (판매수량·판매금액에는 포함)<br>"
    "· 남은 재고 중 가장 오래된 것이 300일 지날 때마다 한 등급씩 강등"
    "</div>", unsafe_allow_html=True)
if mall_sig is None:
    st.warning("몰분류 파일(*몰분류*.xlsx)이 없어 매장(오프라인) 판매를 구분하지 못했습니다. "
               "수익율·이익율에 매장 실적이 섞여 있습니다.")

query = st.text_input("상품 라인명 검색", placeholder="예: A158WA, 590702 ...",
                      help="모델명 부분일치 (대소문자 무시, 공백으로 여러 단어 AND 검색)")

if not query.strip():
    st.info("모델명(라인명)을 입력하면 연도별 평균 수익·수익율과 주문건별 상세가 표시됩니다.")
    st.stop()

# 검색: 고유 모델명(카테고리)에서 먼저 매칭 → 속도/메모리 절약
terms = [t.upper() for t in query.strip().split()]
cats = df[COL_MODEL].cat.categories
cats_up = cats.astype(str).str.upper()
hit_mask = np.ones(len(cats), dtype=bool)
for t in terms:
    hit_mask &= np.asarray(cats_up.str.contains(re.escape(t), na=False, regex=True), dtype=bool)
matched = set(cats[hit_mask])

if not matched:
    st.warning(f"'{query}' 검색 결과가 없습니다.")
    st.stop()

hit = df[df[COL_MODEL].isin(matched)]

brands = [str(b) for b in hit[COL_BRAND].dropna().unique()[:5]]

def 온라인만(sub: pd.DataFrame) -> pd.DataFrame:
    """수익성 지표에서는 매장(오프라인) 판매를 뺀다.

    매장은 수수료·배송비 구조가 온라인과 달라 섞으면 수익율이 왜곡된다.
    반대로 판매수량·판매금액은 실제로 팔린 것이므로 매장도 포함해야 한다.
    """
    return sub[~sub["매장"]] if "매장" in sub.columns else sub


# 상품등급: 이익율(수익 ÷ 정산금) 기준 S~E
# S 는 이익율만으로 주지 않는다. 몇 개 안 팔린 상품이 이익율만 높아
# 최고 등급을 받는 걸 막기 위해 판매수량 조건을 같이 건다.
GRADE_S_RATE = 40      # 이익율(%) 이상이고
GRADE_S_QTY = 20       # 판매수량(개) 이상이면 S
GRADE_CUTS = [(25, "A", "#1a7f37"), (15, "B", "#0969da"),
              (5, "C", "#bf8700"), (-5, "D", "#d4691e")]

def grade_of(sub: pd.DataFrame):
    sub = 온라인만(sub)                    # 등급도 온라인 실적으로만 매긴다
    settle = sub["정산금_수익"].sum()      # 배송비까지 뺀 금액이 분모
    if sub.empty or settle <= 0:
        return None, None
    rate = sub[COL_PROFIT].sum() / settle * 100
    if rate >= GRADE_S_RATE and sub[COL_QTY].sum() >= GRADE_S_QTY:
        return ("S", "#7c3aed"), rate
    for cut, gr, color in GRADE_CUTS:
        if rate >= cut:
            return (gr, color), rate
    return ("E", "#cf222e"), rate

# ── 등급 산출 범위 ────────────────────────────────────
# 최근 3개월부터 보되, 표본이 얇으면 3개월씩 뒤로 늘려간다.
#   3개월 → 100건 넘으면 확정
#   안 되면 6, 9, 12 ... 최대 24개월까지
#   24개월로도 100건이 안 되면 그 24개월치를 그대로 쓴다
# 잘 팔리는 상품은 최근 실적만, 드물게 팔리는 상품은 기간을 넓혀
# 최소한의 표본을 확보하는 방식.
GRADE_TARGET = 100        # 이 건수를 넘기면 기간 확장을 멈춘다
GRADE_STEP_MONTHS = 3     # 늘리는 단위(개월)
GRADE_MAX_MONTHS = 24     # 최대 기간(개월)

TURN_MONTHS = 6      # 회전율 기준 기간(개월). 묵은 재고가 있으면 그만큼 늘어남

_g_pool = (hit[hit["연도"].isin([2024, 2025, 2026])]
           .sort_values(COL_DATE, ascending=False))
# '지금'은 실행 시각이 아니라 데이터의 마지막 출고일로 본다
# (파일 갱신이 하루이틀 늦어도 기준이 흔들리지 않게)
_g_today = df[COL_DATE].max()

_g_base, g_basis = _g_pool, "전체"
for _개월 in range(GRADE_STEP_MONTHS, GRADE_MAX_MONTHS + 1, GRADE_STEP_MONTHS):
    _g_base = _g_pool[_g_pool[COL_DATE] >= _g_today - pd.DateOffset(months=_개월)]
    g_basis = f"최근 {_개월}개월"
    if len(_g_base) > GRADE_TARGET:
        break

# 2년간 판매가 아예 없으면 그 기간으로는 등급을 못 매긴다 → 있는 것 전부로
if _g_base.empty:
    _g_base, g_basis = _g_pool, f"전체 {len(_g_pool):,}건"

g_res, g_rate = grade_of(_g_base)
g_n = len(_g_base)

# ── 재고 매칭 (등급 밑 표시 + 연도별 입고 역산) ──
_RX_IN = re.compile(r"(\d+)일전/(\d+)")
stock_df = load_stock(get_stock_sig())
stock_info, inbound_by_year = None, {}
if stock_df is not None:
    smask = pd.Series(True, index=stock_df.index)
    for t in terms:
        smask &= stock_df["모델명_U"].str.contains(re.escape(t), na=False, regex=True)
    s_hit = stock_df[smask]
    if not s_hit.empty:
        _qty = int(s_hit["수량"].sum())
        _inb_hist = int(s_hit["총입고량"].sum())
        # 입고이력은 최근 3회까지만 기록되어 과소 집계됨
        # → 실제 판매수량 + 현재고로 하한을 보정
        _sold = int(hit[COL_QTY].sum())
        _inb = max(_inb_hist, _sold + _qty)
        stock_info = {
            "라인명들": (sorted(set(s_hit["라인명"].dropna().astype(str)))
                     if "라인명" in s_hit.columns else []),
            # 재고 파일에 같이 들어 있는 EC행사가와 그 수익률
            # (여러 사이즈가 잡히면 대표값 하나만 — 보통 라인 전체가 같은 값)
            "EC행사가": (pd.to_numeric(s_hit["EC행사가"], errors="coerce").dropna().median()
                     if "EC행사가" in s_hit.columns else None),
            "EC행사수익률": (pd.to_numeric(s_hit["EC행사수익률"], errors="coerce").dropna().median()
                        if "EC행사수익률" in s_hit.columns else None),
            "자사몰행사가": (pd.to_numeric(s_hit["자사몰행사가"], errors="coerce").dropna().median()
                       if "자사몰행사가" in s_hit.columns else None),
            "현재고": _qty, "총입고량": _inb, "판매량": _sold,
            "입고추정": _inb > _inb_hist,
            "회전율": (_sold / _inb) if _inb else None,
            "최근입고": s_hit["입고경과일"].min(),
        }
        # 연도별 입고량 역산 + FIFO 기준 "가장 오래된 잔여재고"의 입고 경과일
        # (오래된 입고분부터 판매 소진했다고 보고, 현재고가 걸쳐 있는 입고 회차의 경과일 사용)
        oldest_days = None
        if "입고이력" in s_hit.columns and "기준일" in s_hit.columns:
            for _, r in s_hit.iterrows():
                base = pd.Timestamp(r["기준일"])
                events = [(int(d), int(q)) for d, q in
                          _RX_IN.findall(str(r["입고이력"]).replace(",", ""))]
                for d_, q_ in events:
                    yr_in = (base - pd.Timedelta(days=d_)).year
                    yr_in = max(yr_in, 2024)   # 24년 이전 입고 → 24년 귀속
                    inbound_by_year[yr_in] = inbound_by_year.get(yr_in, 0) + q_
                # FIFO: 이 모델의 잔여재고가 어느 입고 회차부터 남았는지
                remain = int(r["수량"])
                if remain <= 0 or not events:
                    continue
                row_days = None
                for d_, q_ in sorted(events, key=lambda x: x[0]):   # 최신 → 과거 순
                    remain -= q_
                    row_days = d_
                    if remain <= 0:
                        break
                # remain>0이면 입고이력(최대 3회)보다 재고가 많음 → 가장 오래된 회차로
                if row_days is not None:
                    oldest_days = row_days if oldest_days is None else max(oldest_days, row_days)
        stock_info["최초입고"] = oldest_days

        # ── 회전율: 최근 6개월 기준 ────────────────────────────
        # 전 기간으로 재면 몇 년 전에 잘 팔린 실적이 지금 안 팔리는 상태를 가려버린다.
        # 다만 6개월만 딱 끊으면 그 전에 들어와 아직도 안 나간 재고가 분모에서 빠져
        # 오히려 회전율이 좋아 보인다. → 묵은 재고가 남아 있으면 그 입고 시점까지 늘린다.
        _기준일 = pd.Timestamp(s_hit["기준일"].iloc[0]) if "기준일" in s_hit.columns else None
        if _기준일 is None or pd.isna(_기준일):
            _기준일 = df[COL_DATE].max()
        _창일수 = (_기준일 - (_기준일 - pd.DateOffset(months=TURN_MONTHS))).days
        if oldest_days and oldest_days > _창일수:
            _창일수 = int(oldest_days)          # 잔여재고가 들어온 날까지 소급
        _창시작 = _기준일 - pd.Timedelta(days=_창일수)

        _sold_win = int(hit.loc[hit[COL_DATE] >= _창시작, COL_QTY].sum())
        _inb_win = 0
        for _, r in s_hit.iterrows():
            for d_, q_ in _RX_IN.findall(str(r["입고이력"]).replace(",", "")):
                if int(d_) <= _창일수:
                    _inb_win += int(q_)
        # 입고이력은 최근 3회까지만 남아 과소 집계된다.
        # 기간이 잔여재고 입고일까지 걸쳐 있으므로 '기간 판매 + 현재고'가 입고량의 하한.
        _inb_win_adj = max(_inb_win, _sold_win + _qty)
        stock_info.update({
            "회전율": (_sold_win / _inb_win_adj) if _inb_win_adj else None,
            "회전기간일": _창일수,
            "회전기간연장": bool(oldest_days and oldest_days > (
                _기준일 - (_기준일 - pd.DateOffset(months=TURN_MONTHS))).days),
            "기간판매": _sold_win,
            "기간입고": _inb_win_adj,
            "기간입고추정": _inb_win_adj > _inb_win,
        })

        # 완판(현재고 0) 상품: 최초 입고 → 마지막 판매까지 소요 기간
        stock_info["완판일수"] = None
        if stock_info["현재고"] == 0 and "기준일" in s_hit.columns:
            first_in, base_ts = None, None
            for _, r in s_hit.iterrows():
                evs = _RX_IN.findall(str(r["입고이력"]).replace(",", ""))
                if not evs:
                    continue
                b = pd.Timestamp(r["기준일"])
                d_max = max(int(d) for d, _ in evs)
                cand = b - pd.Timedelta(days=d_max)
                if first_in is None or cand < first_in:
                    first_in, base_ts = cand, b
            last_sale = hit[COL_DATE].max() if not hit.empty else None
            if first_in is not None and pd.notna(last_sale):
                stock_info["완판일수"] = max((pd.Timestamp(last_sale) - first_in).days, 0)

# 최초 입고 300일 경과마다 한 등급씩 강등 (600일=2등급, 900일=3등급 ...)
GRADE_ORDER = ["S", "A", "B", "C", "D", "E"]
GRADE_COLORS = {"A": "#1a7f37", "B": "#0969da", "C": "#bf8700", "D": "#d4691e", "E": "#cf222e"}
demote_steps = 0
if (g_res and stock_info and stock_info.get("최초입고") is not None):
    demote_steps = int(stock_info["최초입고"] // 300)
    if demote_steps > 0:
        _i = min(GRADE_ORDER.index(g_res[0]) + demote_steps, len(GRADE_ORDER) - 1)
        demote_steps = _i - GRADE_ORDER.index(g_res[0])   # 실제 적용된 단계
        _g2 = GRADE_ORDER[_i]
        g_res = (_g2, GRADE_COLORS[_g2])

st.markdown(f"**검색 결과 {len(hit):,}건** · 모델 {len(matched):,}종 · "
            f"몰 {hit[COL_MALL].nunique()}곳 · 브랜드: {', '.join(brands)}")


# ── 상단 KPI ──
def agg_stats(sub: pd.DataFrame) -> dict:
    if sub.empty:
        return {"수량": 0, "수익율": None, "평균정산금": None, "이익율": None,
                "매장수량": 0}
    qty = sub[COL_QTY].sum()
    settle = sub["정산금"].sum()          # 배송비 차감 안 한 금액 (표시용)

    # 수익율·이익율은 온라인만
    on = 온라인만(sub)
    sales_on = on[COL_PRICE].sum()
    profit_on = on[COL_PROFIT].sum()
    settle_on = on["정산금_수익"].sum()   # 배송비까지 뺀 금액 (수익 기준)

    return {"수량": int(qty),
            "매장수량": int(sub.loc[sub["매장"], COL_QTY].sum()) if "매장" in sub.columns else 0,
            # 수익율 = 수익 ÷ 판매가   (고객이 낸 돈 대비)
            "수익율": (profit_on / sales_on * 100) if sales_on else None,
            # 이익율 = 수익 ÷ 정산금(배송비 차감)  — 상품등급과 같은 산식
            "이익율": (profit_on / settle_on * 100) if settle_on else None,
            # 평균 정산금은 실제 정산된 금액 기준, 매장 포함
            "평균정산금": (settle / qty) if qty else None}


years = [2024, 2025, 2026]
periods = [("최근 3개년", hit[hit["연도"].isin(years)])]
periods += [(f"{y}년", hit[hit["연도"] == y]) for y in years]

# 26년 칸과 상품등급 칸 사이에 상품 이미지
# 검색된 모델명 → (사이즈 뗀 이름) → 재고에서 찾은 라인명 순으로 이미지를 찾는다
img_map = load_images(get_image_sig())
_img_후보 = list(matched) + list(stock_info.get("라인명들", []) if stock_info else [])
_u = _이미지찾기(img_map, _img_후보)
img_url = _이미지주소(_u) if _u else None

cols = st.columns([1, 1, 1, 1, 0.85, 1.25])
for col, (label, sub) in zip(cols, periods):
    s = agg_stats(sub)
    with col:
        if label == "최근 3개년":
            inb = sum(inbound_by_year.get(y, 0) for y in years)
        else:
            inb = inbound_by_year.get(int(label[:4]), 0)
        inb_txt = f"{inb:,}" if stock_info else "-"
        inb_label = "입고(~24년)" if label == "2024년" else "입고"
        if s["수량"]:
            st.markdown(f"**{label}**")
            st.caption(f"{inb_label} {inb_txt}개 / 판매 {s['수량']:,}개")
            st.metric("평균 수익율", fmt_pct(s["수익율"]))
            # 정산금·이익율은 st.metric 두 개를 나란히 놓으면 칸을 넘쳐 글자가 겹친다.
            # → 한 줄에 작게 (이익율은 상품등급과 같은 산식: 수익 ÷ 정산금)
            _한줄지표("평균 정산금", fmt_won(s["평균정산금"]),
                   "이익율", fmt_pct(s["이익율"]))
        else:
            st.markdown(f"**{label}**")
            st.caption(f"{inb_label} {inb_txt}개 / 판매 0개")
            st.metric("평균 수익율", "-")
            _한줄지표("평균 정산금", "-", "이익율", "-")

with cols[4]:
    st.markdown("**상품 이미지**")
    if img_url:
        try:
            st.image(img_url, width="stretch")
        except Exception:
            st.caption("이미지를 불러오지 못했습니다")
    elif img_map:
        st.caption("이미지 없음")
    else:
        st.caption("이미지 파일 없음")

with cols[5]:
    # 등급은 최근 실적만 본다. 표본이 얇으면 기간을 3개월씩 늘려 잡는다.
    st.markdown(f"**상품등급** <span style='font-size:0.78rem;color:#888;font-weight:400'>"
                f"({g_basis} 기준)</span>", unsafe_allow_html=True)
    _ec율 = (stock_info or {}).get("EC행사수익률")
    if _ec율 is not None and (pd.isna(_ec율) or abs(_ec율) > 5):
        _ec율 = None                      # 이상값(수천 %) 은 표시하지 않는다
    if g_res:
        st.markdown(
            f"<div style='line-height:1.05;margin-bottom:0.4rem'>"
            f"<span style='font-size:2.6rem;font-weight:900;color:{g_res[1]}'>{g_res[0]}</span>"
            f" <span style='font-size:0.8rem;color:#666'>이익율 {g_rate:.2f}%"
            + (f" <span style='color:#999'>({g_n:,}건)</span>" if g_n <= GRADE_TARGET else "")
            # 재고 파일 L열의 EC행사 수익률. 실제 판매 실적이 아니라
            # 행사가 기준으로 계산된 값이라 이익율과 나란히 두고 비교한다.
            + (f"<br>EC기준 수익율 <b style='color:#333'>{_ec율 * 100:.1f}%</b>"
               if _ec율 is not None else "")
            + "</span></div>",
            unsafe_allow_html=True)
    else:
        st.caption("등급 산출 불가")

    # 가장 오래된 잔여재고의 입고 경과일 — 강등 여부와 무관하게 항상 보여준다
    # (등급이 좋아도 재고가 묵고 있으면 바로 눈에 띄어야 하므로)
    if stock_info:
        _old = stock_info.get("최초입고")
        if _old is None:
            _old_txt = "<span style='color:#999'>잔여재고 없음</span>"
        else:
            _old_col = "#cf222e" if _old >= 300 else ("#bf8700" if _old >= 180 else "#666")
            _old_txt = (f"<span style='color:{_old_col};font-weight:700'>{int(_old):,}일</span> 경과"
                        + (f" <span style='color:#cf222e'>⬇ -{demote_steps}등급</span>"
                           if demote_steps else ""))
        st.markdown(
            "<div style='font-size:0.8rem;color:#666;margin-top:-0.25rem;margin-bottom:0.35rem'>"
            f"가장 오래된 재고 {_old_txt}</div>", unsafe_allow_html=True)

    if stock_info:
        _t = stock_info["회전율"]
        _r = stock_info["최근입고"]
        _turn_label = (f"최근 {TURN_MONTHS}개월"
                       if not stock_info.get("회전기간연장")
                       else f"최근 {stock_info['회전기간일'] / 30:.0f}개월 · 묵은재고 입고시점까지")
        st.markdown(
            "<div style='font-size:0.85rem;line-height:1.7;border-top:1px solid #e6e6e6;padding-top:0.35rem'>"
            f"총입고 <b>{stock_info['총입고량']:,}개</b>"
            + ("<span style='color:#999'>*</span>" if stock_info.get("입고추정") else "")
            + "<br>"
            f"판매 <b>{stock_info['판매량']:,}개</b><br>"
            f"현재고 <b>{stock_info['현재고']:,}개</b><br>"
            + (f"회전율(판매÷입고) <b>{_t:.1%}</b>" if _t is not None and pd.notna(_t)
               else "회전율(판매÷입고) <b>-</b>")
            + f" <span style='color:#999;font-size:0.75rem'>({_turn_label})</span><br>"
            + f"<span style='color:#999;font-size:0.75rem'>└ 기간 판매 {stock_info['기간판매']:,}"
              f" ÷ 입고 {stock_info['기간입고']:,}</span><br>"
            + (f"최근입고 <b>{int(_r)}일 전</b>" if pd.notna(_r) else "최근입고 <b>-</b>")
            + (f"<br>✅ 완판까지 <b>{stock_info['완판일수']:,}일</b>"
               if stock_info.get("완판일수") is not None else "")
            + _행사가표시(stock_info)
            + ("<br><span style='color:#999;font-size:0.75rem'>* 입고이력 누락분 보정(판매+재고)</span>"
               if stock_info.get("입고추정") else "")
            + "</div>", unsafe_allow_html=True)
    else:
        st.caption("재고 데이터 없음")

st.divider()

# ── 수익율 구간별 판매수량 ──
st.subheader("수익율 구간별 판매수량")

BIN_LO, BIN_HI, BIN_STEP = -5, 50, 5
edges = list(range(BIN_LO, BIN_HI + BIN_STEP, BIN_STEP))
labels = [f"{BIN_LO}% 미만"] + [f"{a}~{b}%" for a, b in zip(edges[:-1], edges[1:])] + [f"{BIN_HI}% 이상"]
bins = [-np.inf] + edges + [np.inf]

hit_b = hit.assign(구간=pd.cut(hit["수익율"], bins=bins, labels=labels, right=False))

# 전체 기준으로 사용할 구간 범위 결정 (연도 간 x축 통일)
tot = hit_b.groupby("구간", observed=False)[COL_QTY].sum()
tot = tot[tot.cumsum() > 0]
tot = tot[::-1][(tot[::-1].cumsum() > 0)][::-1]
used = list(tot.index)

# 연도 간 y축 통일: 전체 연도에서 구간별 최대 판매수량
y_max = int((hit_b[hit_b["구간"].isin(used)]
             .groupby(["연도", "구간"], observed=True)[COL_QTY].sum().max() or 0))

chart_cols = st.columns(len(years))
for col, yr in zip(chart_cols, years):
    with col:
        sub = hit_b[(hit_b["연도"] == yr) & (hit_b["구간"].isin(used))]
        st.markdown(f"**{yr}년**")
        if sub.empty:
            st.caption("데이터 없음")
            continue

        agg = (sub.groupby("구간", observed=False)
               .agg(판매수량=(COL_QTY, "sum"), 정산금합=("정산금", "sum"))
               .reindex(used))
        agg["평균정산금"] = agg["정산금합"] / agg["판매수량"]
        agg["판매수량"] = agg["판매수량"].fillna(0).astype(int)
        agg = agg[agg["판매수량"] > 0]          # 판매 없는 구간 제거

        # x축 라벨: 구간 + (해당 연도 평균 정산금, 만원 축약) 2줄 표시
        def _lab(b, v):
            return f"{b}|({v/10000:,.1f}만)" if pd.notna(v) else f"{b}|(-)"
        chart_df = agg.reset_index()
        chart_df["구간라벨"] = [_lab(b, v) for b, v in zip(chart_df["구간"].astype(str),
                                                           chart_df["평균정산금"])]
        chart_df["평균정산금"] = chart_df["평균정산금"].round(0)

        st.altair_chart(
            alt.Chart(chart_df).mark_bar().encode(
                x=alt.X("구간라벨:N", sort=list(chart_df["구간라벨"]),
                        title=None,
                        axis=alt.Axis(labelAngle=0, labelExpr="split(datum.label, '|')",
                                      labelFontSize=11, labelFontWeight="bold",
                                      labelColor="#31333F", labelLimit=0,
                                      labelOverlap=False)),
                y=alt.Y("판매수량:Q", title=None,
                        scale=alt.Scale(domain=[0, y_max])),
                tooltip=[alt.Tooltip("구간:N", title="수익율 구간"),
                         alt.Tooltip("판매수량:Q", format=","),
                         alt.Tooltip("평균정산금:Q", format=",.0f", title="평균 정산금(원)")],
            ).properties(height=260),
            use_container_width=True,
        )

st.divider()

# ── 몰별 상세 ──
st.subheader("몰별 상세")


def mall_summary(sub: pd.DataFrame) -> pd.DataFrame:
    # 몰별 상세는 매장도 포함해서 보여준다 (실제 판매 내역이므로)
    g = (sub.groupby(COL_MALL, observed=True)
         .agg(수량합=(COL_QTY, "sum"), 총매출=(COL_PRICE, "sum"),
              _정산금합=("정산금", "sum"), _수익합=(COL_PROFIT, "sum"))
         .sort_values("총매출", ascending=False))
    g["평균정산금"] = (g["_정산금합"] / g["수량합"]).round(0)
    g["평균수익율(%)"] = (g["_수익합"] / g["총매출"] * 100).round(2)
    return g.drop(columns=["_정산금합", "_수익합"]).reset_index()


sum_cols = st.columns(len(years))
for col, yr in zip(sum_cols, years):
    with col:
        st.markdown(f"**{yr}년**")
        sub_y = hit[hit["연도"] == yr]
        if sub_y.empty:
            st.caption("데이터 없음")
            continue
        st.dataframe(
            mall_summary(sub_y), hide_index=True,
            column_config={"수량합": NUM, "총매출": WON, "평균정산금": WON,
                           "평균수익율(%)": st.column_config.NumberColumn(format="%.2f%%")},
        )


st.divider()

# ── 주문건별 상세 ──
st.subheader("주문건별 상세")

def 등급범위(pool: pd.DataFrame) -> pd.DataFrame:
    """등급 산출에 쓸 기간을 고른다 (위 상품등급과 같은 규칙)."""
    pool = pool.sort_values(COL_DATE, ascending=False)
    base = pool
    for 개월 in range(GRADE_STEP_MONTHS, GRADE_MAX_MONTHS + 1, GRADE_STEP_MONTHS):
        base = pool[pool[COL_DATE] >= _g_today - pd.DateOffset(months=개월)]
        if len(base) > GRADE_TARGET:
            break
    return pool if base.empty else base


f1, f2, f3 = st.columns([2, 2, 1])
with f1:
    mall_sel = st.multiselect("쇼핑몰", sorted(str(x) for x in hit[COL_MALL].dropna().unique()))
with f2:
    model_sel = st.multiselect("모델명", sorted(str(x) for x in matched))
with f3:
    year_sel = st.multiselect("연도", sorted(hit["연도"].unique(), reverse=True))

detail = hit
if mall_sel:
    detail = detail[detail[COL_MALL].astype(str).isin(mall_sel)]
if model_sel:
    detail = detail[detail[COL_MODEL].astype(str).isin(model_sel)]
if year_sel:
    detail = detail[detail["연도"].isin(year_sel)]

detail = detail.sort_values(COL_DATE, ascending=False)

show = detail[[COL_DATE, COL_MALL, COL_BRAND, COL_MODEL, COL_QTY, COL_COST,
               COL_PRICE, "정산금", COL_PROFIT, "수익율", COL_NOTE]].rename(
    columns={COL_COST: "원가", COL_PRICE: "최종판매가",
             COL_PROFIT: "수익(실배송비)", "수익율": "수익율(%)"})
show[COL_DATE] = show[COL_DATE].dt.strftime("%Y-%m-%d")
# 어느 모델이 무슨 등급인지 표에서도 바로 보이게

st.dataframe(
    show,
    hide_index=True,
    height=520,
    column_config={"수량": NUM, "원가": WON, "최종판매가": WON, "정산금": WON,
                   "수익(실배송비)": WON,
                   "수익율(%)": st.column_config.NumberColumn(format="%.2f%%")},
)

st.download_button("CSV 다운로드", show.to_csv(index=False).encode("utf-8-sig"),
                   file_name=f"수익율_{query.strip()}.csv", mime="text/csv")

st.divider()

# ──────────────────────────────────────────────
# 전체 상품 등급 내려받기
# ──────────────────────────────────────────────
st.subheader("🏅 전체 상품 등급")


@st.cache_data(show_spinner="전체 등급 계산 중...")
def 전체등급표(file_sigs, stock_sig, mall_sig) -> pd.DataFrame:
    """모든 라인명의 등급을 한 번에 계산한다.

    라인명 하나씩 기간을 넓혀 가며 재면 2만 번을 반복해야 해서 몇 분씩 걸린다.
    → 각 행이 '몇 번째 기간 구간'에 드는지 미리 표시하고,
      구간별 합계를 누적해서 모든 라인명의 기간을 한 번에 정한다.
    """
    d = load_all_data(file_sigs, mall_sig)
    d = 온라인만(d)                                   # 등급은 온라인 실적만
    d = d[d["연도"].isin([2024, 2025, 2026])]
    d = d[d[COL_QTY] > 0]                            # 수량 0 건 제외
    if COL_BRAND in d.columns:                       # 리퍼는 등급 대상이 아니다
        d = d[d[COL_BRAND].astype(str).str.strip() != "리퍼"]
    if d.empty:
        return pd.DataFrame()

    # 모델명 → 라인명 (재고 파일 기준, 없으면 사이즈만 떼어 씀)
    s = load_stock(stock_sig)
    맵 = {}
    if s is not None and "라인명" in s.columns:
        맵 = dict(zip(s["모델명"].astype(str).str.strip().str.upper(),
                     s["라인명"].astype(str).str.strip()))
    모델 = d[COL_MODEL].astype(str)
    라인 = 모델.str.strip().str.upper().map(맵)
    라인 = 라인.fillna(모델.str.replace(_RX_SIZE, "", regex=True).str.strip())
    d = d.assign(라인명=라인)

    # 기간 구간: 최근 3,6,...,24개월. 그 밖은 마지막 칸.
    끝 = d[COL_DATE].max()
    경계 = [끝 - pd.DateOffset(months=k)
          for k in range(GRADE_STEP_MONTHS, GRADE_MAX_MONTHS + 1, GRADE_STEP_MONTHS)]
    구간 = np.searchsorted(np.array(경계[::-1], dtype="datetime64[ns]"),
                         d[COL_DATE].values.astype("datetime64[ns]"), side="right")
    d = d.assign(구간=len(경계) - 구간)                # 0 = 최근 3개월
    d["구간"] = d["구간"].clip(0, len(경계))

    브랜드표 = (d.groupby("라인명", observed=True)[COL_BRAND]
              .agg(lambda x: x.astype(str).mode().iat[0] if len(x) else "")
              if COL_BRAND in d.columns else None)

    합 = (d.groupby(["라인명", "구간"], observed=True)
          .agg(건수=(COL_QTY, "size"), 수량=(COL_QTY, "sum"),
               수익=(COL_PROFIT, "sum"), 정산=("정산금_수익", "sum")))
    # 구간별 누적합. (groupby(axis=1) 은 pandas 최신판에서 없어져 항목별로 따로 돌린다)
    누적 = {이름: 합[이름].unstack("구간", fill_value=0).sort_index(axis=1).cumsum(axis=1)
          for 이름 in ("건수", "수량", "수익", "정산")}

    건수 = 누적["건수"]
    # 100건을 넘기는 첫 구간. 끝까지 못 넘기면 마지막 구간(=24개월, 없으면 전체)
    넘김 = 건수.gt(GRADE_TARGET)
    고른칸 = np.where(넘김.any(axis=1), 넘김.values.argmax(axis=1), 건수.shape[1] - 1)
    행 = np.arange(len(건수))

    표 = pd.DataFrame({
        "라인명": 건수.index.astype(str),
        "건수": 건수.values[행, 고른칸],
        "수량": 누적["수량"].values[행, 고른칸],
        "수익": 누적["수익"].values[행, 고른칸],
        "정산금": 누적["정산"].values[행, 고른칸],
    })
    표["기준기간"] = [f"최근 {(k + 1) * GRADE_STEP_MONTHS}개월"
                  if k < len(경계) else "전체" for k in 고른칸]
    표["이익율(%)"] = np.where(표["정산금"] > 0, 표["수익"] / 표["정산금"] * 100, np.nan)

    def _등급(r):
        v = r["이익율(%)"]
        if pd.isna(v) or r["정산금"] <= 0:
            return "-"
        if v >= GRADE_S_RATE and r["수량"] >= GRADE_S_QTY:
            return "S"
        for cut, gr, _ in GRADE_CUTS:
            if v >= cut:
                return gr
        return "E"

    표["등급"] = 표.apply(_등급, axis=1)
    표["이익율(%)"] = 표["이익율(%)"].round(2)
    표["브랜드"] = (표["라인명"].map(브랜드표) if 브랜드표 is not None else "")
    표 = 표[표["수량"] > 0]                            # 기간 내 판매가 없으면 뺀다
    표 = 표[["라인명", "브랜드", "등급", "이익율(%)", "수량", "건수", "기준기간"]]
    표["_ord"] = 표["등급"].map({g: i for i, g in enumerate(GRADE_ORDER)}).fillna(99)
    return (표.sort_values(["_ord", "수량"], ascending=[True, False])
            .drop(columns=["_ord"]).reset_index(drop=True))


_등급표전체 = 전체등급표(file_sigs, get_stock_sig(), mall_sig)

if _등급표전체.empty:
    st.caption("등급을 낼 수 있는 데이터가 없습니다.")
else:
    _분포 = _등급표전체["등급"].value_counts()
    st.caption("라인명 **{:,}개** · ".format(len(_등급표전체))
               + " · ".join(f"{g} {int(_분포.get(g, 0)):,}" for g in GRADE_ORDER
                            if _분포.get(g, 0))
               + " · 매장(오프라인) 제외, 상단 규칙과 같은 방식")
    st.download_button(
        f"⬇ 전체 상품 등급 {len(_등급표전체):,}개 내려받기 (엑셀)",
        data=_엑셀바이트(_등급표전체, "상품등급"),
        file_name=f"상품등급_{pd.Timestamp.today():%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.divider()

# ──────────────────────────────────────────────
# 이미지 없는 라인명 내려받기
# ──────────────────────────────────────────────
st.subheader("🖼 이미지 없는 라인명")


# 사은품·쇼핑백·PR(홍보용)·리퍼·노다는 원래 상품 이미지가 없는 것이 정상이라 목록에서 뺀다
IMG_SKIP_BRANDS = {"사은품", "쇼핑백", "PR", "리퍼", "노다"}

@st.cache_data(show_spinner=False)
def 이미지없는_라인명(stock_sig, image_sig) -> pd.DataFrame:
    """재고에는 있는데 이미지 표에 모델명이 하나도 없는 라인명."""
    s = load_stock(stock_sig)
    m = load_images(image_sig)
    if s is None or not m or "라인명" not in s.columns:
        return pd.DataFrame()

    s = s.copy()
    if "브랜드" in s.columns:
        s = s[~s["브랜드"].astype(str).str.strip().isin(IMG_SKIP_BRANDS)]
    # 이미지 표는 라인명(사이즈 없는 이름) 기준이라 모델명만으로 찾으면 대부분 놓친다.
    # 모델명 · 사이즈 뗀 모델명 · 라인명 중 하나라도 있으면 '이미지 있음'.
    키 = set(m.keys())
    _m = s["모델명"].astype(str).str.strip().str.upper()
    _l = s["라인명"].astype(str).str.strip().str.upper()
    _b = _m.str.replace(_RX_SIZE, "", regex=True).str.strip()
    s["_있음"] = _m.isin(키) | _b.isin(키) | _l.isin(키)

    있는열 = [c for c in ("브랜드", "대카테고리", "카테고리", "종류", "성별") if c in s.columns]
    집계 = {"모델수": ("모델명", "nunique"),
           "이미지있는모델": ("_있음", "sum"),
           "재고수량": ("수량", "sum"),
           "가용수량": ("가용수량", "sum"),
           "모델명": ("모델명", lambda x: ", ".join(sorted(set(map(str, x)))))}
    for c in 있는열:
        집계[c] = (c, "first")

    g = (s.dropna(subset=["라인명"])
         .groupby("라인명", observed=True).agg(**집계).reset_index())

    없음 = g[g["이미지있는모델"] == 0].drop(columns=["이미지있는모델"])
    순서 = ["라인명"] + 있는열 + ["모델수", "재고수량", "가용수량", "모델명"]
    return (없음[순서]
            .sort_values(["가용수량", "라인명"], ascending=[False, True])
            .reset_index(drop=True))


_없음 = 이미지없는_라인명(get_stock_sig(), get_image_sig())

if _없음.empty:
    if img_map is None:
        st.caption("이미지 파일이 없어 확인할 수 없습니다. (이미지.parquet 또는 이미지.xlsx 필요)")
    elif stock_df is None:
        st.caption("재고 파일이 없어 확인할 수 없습니다.")
    else:
        st.success("이미지가 없는 라인명이 없습니다.")
else:
    _재고있음 = int((_없음["가용수량"] > 0).sum())
    st.caption(f"재고 기준 **{len(_없음):,}개** 라인명에 이미지가 없습니다 "
               f"(그중 가용재고 있는 것 {_재고있음:,}개). 가용수량 많은 순으로 정렬했습니다. "
               f"· {' / '.join(sorted(IMG_SKIP_BRANDS))}는 제외")

    _날짜 = pd.Timestamp.today().strftime("%Y%m%d")
    if stock_df is not None and "기준일" in stock_df.columns and len(stock_df):
        try:
            _날짜 = pd.Timestamp(stock_df["기준일"].iloc[0]).strftime("%Y%m%d")
        except Exception:
            pass

    # CSV 대신 xlsx. 코드성 컬럼을 '텍스트' 서식으로 넣어야
    # 710548524014 같은 값이 지수형으로 바뀌거나 앞자리 0 이 날아가지 않는다.
    st.download_button(
        f"⬇ 이미지 없는 라인명 {len(_없음):,}개 내려받기 (엑셀)",
        data=_엑셀바이트(_없음, "이미지없는라인명"),
        file_name=f"이미지없는_라인명_{_날짜}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
